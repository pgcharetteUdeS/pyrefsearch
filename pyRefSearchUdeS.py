"""pyRefSearchUdeS.py

    For a list of author names and range of years supplied in an input Excel file,
    query references (publications in Scopus, patents in the USPTO database) OR
    author profiles (Scopus database), and write the results to an output Excel file.

    All execution parameters specified in the file "pyRefSearchUdeS.toml"

    The script uses the "pybliometrics" for Scopus searches,
    see https://pybliometrics.readthedocs.io/en/stable/
    NB: A "key" is required to query the Scopus API,
        see https://dev.elsevier.com/index.jsp. The first execution of the script
        will prompt the user to enter the key.

    The script uses the "patent_client" package for searches in the USPTO database,
    see https://patent-client.readthedocs.io/en/latest/user_guide/fulltext.html
    NB: support for European Patent Office (EPO) searches is pending

    Project on gitHub: https://github.com/pgcharetteUdeS/pyRefSearchUdeS

"""

from functools import lru_cache
import numpy as np
from openpyxl import load_workbook
import pybliometrics
import pandas as pd
from patent_client import Inpadoc, Patent, PublishedApplication
from pathlib import Path
from pybliometrics.scopus import AuthorRetrieval, AuthorSearch, ScopusSearch
from pybliometrics.scopus.exception import ScopusException
import re
from rich import print
import sys
import toml
import unidecode
from version import __version__
import warnings


class ReferenceQuery:
    """
    Class to store reference query parameters
    """

    @staticmethod
    def check_excel_file_access(filename: Path):
        try:
            with open(filename, "a"):
                pass
        except IOError:
            raise IOError(
                f"Could not open '{filename}', close it " "if it's already open!"
            ) from None

    def __init__(
        self,
        in_excel_file: Path,
        in_excel_file_author_sheet: str,
        out_excel_file: Path,
        pub_year_first: int,
        pub_year_last: int,
        publication_types: list[str],
        local_affiliations: list[str],
        scopus_database_refresh: bool | int,
        include_patents: bool = True,
    ):
        self.in_excel_file: Path = in_excel_file
        self.out_excel_file: Path = out_excel_file
        self.pub_year_first: int = pub_year_first
        self.pub_year_last: int = pub_year_last
        self.publication_types: list[str] = [row[0] for row in publication_types]
        self.publication_type_codes: list[str] = [row[1] for row in publication_types]
        self.local_affiliations: list[str] = [
            _to_lower_no_accents_no_hyphens(s) for s in local_affiliations
        ]
        self.scopus_database_refresh: bool | int = scopus_database_refresh
        self.include_patents: bool = include_patents

        # Check input/output Excel file access, script fails if files already open
        self.check_excel_file_access(self.in_excel_file)
        self.check_excel_file_access(self.out_excel_file)

        # Load input Excel file into a dataframe, remove rows without author names
        warnings.simplefilter(action="ignore", category=UserWarning)
        input_data_full: pd.DataFrame = pd.read_excel(
            self.in_excel_file, sheet_name=in_excel_file_author_sheet
        )
        input_data_full = input_data_full.dropna(subset=["Nom"])

        # Extract author names from input Excel file, formatted either as a 3IT database
        # (author status tabulated by fiscal year) or as a simple list of names
        author_status_by_year_columns: list[str] = [
            f"{year}-{year + 1}"
            for year in range(self.pub_year_first, self.pub_year_last + 1)
        ]
        if all(col in input_data_full.columns for col in author_status_by_year_columns):
            # Author information is tabulated by fiscal year (XXXX-YYYY) and status (full
            # member or collaborator). Validate that the range of years specified
            # in the input data covers the range of years specified in the query,
            # filter by member status/year to remove collaborators.
            authors: pd.DataFrame = input_data_full.copy()[
                ["Nom", "Prénom", "ID Scopus"] + author_status_by_year_columns
            ]
            authors["status"] = [
                "Régulier" if "Régulier" in yearly_status else "Collaborateur"
                for yearly_status in authors[
                    author_status_by_year_columns
                ].values.tolist()
            ]
            authors.drop(authors[authors.status == "Collaborateur"].index, inplace=True)

        elif not any(
            # Author information is supplied as a simple list of names, no filtering
            re.search(r"\d{4}-\d{4}", column)
            for column in input_data_full.columns.tolist()
        ):
            authors: pd.DataFrame = input_data_full.copy()[
                ["Nom", "Prénom", "ID Scopus"]
            ]

        else:
            raise IOError(
                f"Range of years [{self.pub_year_first}-{self.pub_year_last}] exceeds "
                f"the available data in '{in_excel_file}'!"
            ) from None
        self.au_names = authors[["Nom", "Prénom"]].values.tolist()

        # Extract Scopus IDs, replace non-integer values with 0
        self.au_ids = []
        if "ID Scopus" in authors:
            for scopus_id in authors["ID Scopus"].values.tolist():
                try:
                    self.au_ids.append(int(scopus_id))
                except ValueError:
                    self.au_ids.append(0)


@lru_cache(maxsize=1024)
def _to_lower_no_accents_no_hyphens(s: str | pd.Series) -> str:
    """
    Convert string to lower case and remove accents and hyphens

    Args:
        s (str): Input string

    Returns: String in lower case without accents

    """

    return unidecode.unidecode(s.replace("-", " ").lower().strip())


def _reindex_author_profiles_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reindex and re-order columns in author profiles DataFrame
    (the Scopus database indexing puts the names far down the list)

    Args:
        df (pd.DataFrame): DataFrame with author profiles

    Returns: DataFrame with re-indexed & re-ordered author profiles

    """

    df.reset_index(drop=True, inplace=True)
    df = df[
        pd.Index(
            [
                "surname",
                "givenname",
                "initials",
                "Affl/ID",
                "Start",
                "End",
                "eid",
                "affiliation",
                "affiliation_id",
                "country",
                "city",
                "orcid",
                "areas",
                "documents",
            ]
        )
    ]
    return df


def _flag_matched_scopus_author_ids_and_affiliations(
    reference_query: ReferenceQuery, author_profiles: pd.DataFrame
) -> pd.DataFrame:
    """
    Flag author profiles with local affiliations and matching Scopus IDs between
    input Excel file and Scopus database

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        author_profiles (pd.DataFrame): DataFrame with author profiles

    Returns: DataFrame with local author profiles flagged

    """

    def set_affiliation_and_id(row):
        if row.affiliation is not None:
            local_affiliation_match: bool = any(
                s in _to_lower_no_accents_no_hyphens(row.affiliation)
                for s in reference_query.local_affiliations
            )
            au_id_index: int | None = (
                reference_query.au_ids.index(int(row.eid))
                if int(row.eid) in reference_query.au_ids
                else None
            )
            au_id_match: bool = (
                au_id_index is not None
                and _to_lower_no_accents_no_hyphens(
                    reference_query.au_names[au_id_index][0]
                )
                == _to_lower_no_accents_no_hyphens(row.surname)
            )
            if local_affiliation_match and au_id_match:
                return "Affl. + ID"
            elif local_affiliation_match:
                return "Affl."
            elif au_id_match:
                return "ID"
            else:
                return None

    author_profiles["Affl/ID"] = author_profiles.apply(set_affiliation_and_id, axis=1)

    return author_profiles


def _check_author_name_correspondance(
    reference_query: ReferenceQuery, authors: pd.DataFrame
) -> list:
    """

    Check that the author names & affiliations supplied in the input Excel file
    correspond to the information associated with their IDs in the Scopus database

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        authors (pd.DataFrame): DataFrame with author profiles

    Returns: List of errors per author

    """

    # Loop through authors to check for discrepancies/errors
    query_errors: list[str | None] = []
    for i, [
        [input_last_name, input_first_name],
        au_id,
        scopus_last_name,
        scopus_first_name,
        affiliation,
        parent_affiliation,
    ] in enumerate(
        zip(
            reference_query.au_names,
            reference_query.au_ids,
            authors["Nom de famille"],
            authors["Prénom"],
            authors["Affiliation"],
            authors["Affiliation mère"],
        )
    ):
        query_error: str | None = None
        if scopus_last_name is None:
            # Missing Scopus ID, enter name manually into authors profile dataframe
            authors.loc[i, "Nom de famille"] = input_last_name
            authors.loc[i, "Prénom"] = input_first_name
            query_error = "Aucun identifiant Scopus"
            print(
                f"[yellow]WARNING: L'auteur.e '{input_last_name}, {input_first_name}' "
                "n'a pas d'identifiant Scopus[/yellow]"
            )
        else:
            # Check for name discrepancies between input and Scopus database
            scopus_last_name: str = _to_lower_no_accents_no_hyphens(scopus_last_name)
            if scopus_last_name != _to_lower_no_accents_no_hyphens(input_last_name):
                query_error = "Disparité de noms de famille"
                print(
                    f"[red]ERREUR pour l'identifiant {au_id}: "
                    f"le nom de famille de l'auteur.e '{input_last_name}, "
                    f"{input_first_name}' dans {reference_query.in_excel_file} diffère"
                    f" de '{scopus_last_name}, {scopus_first_name}'"
                    " dans la base de données Scopus![/red]"
                )

            # Check for affiliation discrepancies between input and Scopus database
            affiliation: str = (
                ""
                if affiliation is None
                else _to_lower_no_accents_no_hyphens(affiliation)
            )
            parent_affiliation: str = (
                ""
                if parent_affiliation is None
                else _to_lower_no_accents_no_hyphens(parent_affiliation)
            )
            if all(
                s not in affiliation and s not in parent_affiliation
                for s in reference_query.local_affiliations
            ):
                query_error = (
                    "Affiliation non locale"
                    if query_error is None
                    else "Disparité de noms de famille / Affiliation non locale"
                )
                print(
                    f"[red]ERREUR pour l'identifiant {au_id} "
                    f"({input_last_name}, {input_first_name}): "
                    f"l'affiliation '{affiliation}, {parent_affiliation}' "
                    "est non locale![/red]"
                )

        # Append current error to author error list
        query_errors.append(query_error)

    return query_errors


def _count_publications_by_type_in_df(
    reference_query: ReferenceQuery, df: pd.DataFrame
) -> list:
    """
    Count number of publications by type in a dataframe

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        df (pd.DataFrame): DataFrame with publications

    Returns: List of counts per publication type

    """

    if df.empty:
        return [None] * len(reference_query.publication_type_codes)
    else:
        return [
            (
                len(df[df["subtype"] == pub_type])
                if len(df[df["subtype"] == pub_type]) > 0
                else None
            )
            for pub_type in reference_query.publication_type_codes
        ]


def _export_publications_df_to_excel_sheet(
    writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str
) -> None:
    """
    Write selected set of publication dataframe columns to Excel sheet

    Args:
        writer (pd.ExcelWriter): openpyxl writer object
        df (pd.DataFrame): articles dataframe
        sheet_name (str): Excel file sheet name

    Returns: None

    """

    if not df.empty:
        df_copy: pd.DataFrame = df.rename(
            columns={
                "coverDate": "Date",
                "title": "Titre",
                "Nb co-auteurs locaux": "Nb co-auteurs locaux",
                "Auteurs locaux": "Auteurs locaux",
                "author_names": "Auteurs",
                "publicationName": "Publication",
                "volume": "Volume",
                "pageRange": "Pages",
                "doi": "DOI",
            },
        ).copy()
        df_copy[
            [
                "Titre",
                "Date",
                "Nb co-auteurs locaux",
                "Auteurs locaux",
                "Auteurs",
                "Publication",
                "Volume",
                "Pages",
                "DOI",
            ]
        ].to_excel(writer, index=False, sheet_name=sheet_name, freeze_panes=(1, 1))


def _tabulate_patents_per_author(
    reference_query: ReferenceQuery,
    patents: pd.DataFrame,
) -> list:
    """
    Tabulate number of patents or patent applications per author

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        patents (pd.DataFrame): patent search results

    Returns: Number of patents or patent applications per author (list)

    """

    if patents.empty:
        return [None] * len(reference_query.au_ids)

    def inventor_match(inventors) -> bool:
        return any(
            _to_lower_no_accents_no_hyphens(lastname)
            in _to_lower_no_accents_no_hyphens(inventor)
            and _to_lower_no_accents_no_hyphens(firstname)
            in _to_lower_no_accents_no_hyphens(inventor)
            for inventor in inventors
        )

    author_patent_counts: list[int | None] = []
    for [lastname, firstname] in reference_query.au_names:
        count: int = sum(patents["Inventeurs"].apply(inventor_match))
        author_patent_counts.append(count if count > 0 else None)

    return author_patent_counts


def _add_coauthor_columns_and_clean_up_publications_df(
    publications_in: pd.DataFrame, reference_query: ReferenceQuery
) -> pd.DataFrame:
    """
    Add columns listing names and counts of local coauthors to the publications DataFrame,
    and sort by publication date

    Args:
        publications_in (pd.DataFrame): DataFrame with publications
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info

    Returns: DataFrame with added columns and sorted by publication date
    """

    # Remove duplicates
    publications: pd.DataFrame = publications_in.drop_duplicates("eid").copy()

    # Add columns listing names and counts of local coauthors
    def find_local_coauthors(author_ids) -> list:
        co_authors_local: list[str] = [
            name[0]
            for name, au_id in zip(reference_query.au_names, reference_query.au_ids)
            if any([str(au_id) in author_ids]) and au_id > 0
        ]
        return co_authors_local

    publications["Auteurs locaux"] = publications["author_ids"].apply(
        find_local_coauthors
    )
    publications["Nb co-auteurs locaux"] = [
        len(co_authors) if len(co_authors) > 1 else None
        for co_authors in publications["Auteurs locaux"]
    ]

    # Sort by publication date
    publications = publications.sort_values(by=["coverDate"])

    return publications


def _query_uspto(
    reference_query: ReferenceQuery, applications: bool = True
) -> pd.DataFrame:
    """
    Query the USPTO database for patent applications or published patents

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        applications (bool): Search filed applications if True, else search published patents if False

    Returns: DataFrame with patent search results

    """

    def inventor_query_str(inventor: list[str]) -> str:
        return f"({inventor[1]} NEAR2 {inventor[0]})"

    def build_uspto_patent_query_string(field_code: str) -> str:
        s: str = (
            f'@{field_code}>="{reference_query.pub_year_first}0101"'
            f'<="{reference_query.pub_year_last}1231" AND ('
        )
        s += inventor_query_str(reference_query.au_names[0])
        for name in reference_query.au_names[1:]:
            s += " OR "
            s += inventor_query_str(name)
        s += ")"
        return s

    max_results: int = 500
    if applications:
        query_str: str = build_uspto_patent_query_string(field_code="AD")
        patents: pd.DataFrame = (
            PublishedApplication.objects.filter(query=query_str)
            .limit(max_results)
            .values(
                "app_filing_date",
                "guid",
                "appl_id",
                "patent_title",
                "inventors",
                "assignees",
                "related_apps",
            )
            .to_pandas()
        )

    else:
        query_str: str = build_uspto_patent_query_string(field_code="PD")
        patents: pd.DataFrame = (
            Patent.objects.filter(query=query_str)
            .limit(max_results)
            .values(
                "publication_date",
                "app_filing_date",
                "guid",
                "appl_id",
                "patent_title",
                "inventors",
                "assignees",
                "related_apps",
            )
            .to_pandas()
        )

    patents["appl_id"] = patents["appl_id"].astype(int)
    return patents


def _reformat_uspto_search_results(
    patents: pd.DataFrame, applications: bool
) -> pd.DataFrame:
    """
    Reorder USPTO search results by filing date for applications or publication date for
    patents, change column names to French, and remove unnecessary columns

    Args:
        applications (bool): Search filed applications if True, else search published patents if False
        patents (pd.DataFrame): DataFrame with patent search results

    Returns: DataFrame with reordered search results

    """
    if applications:
        patents.rename(
            columns={
                "app_filing_date": "Date de dépôt",
                "guid": "GUID",
                "appl_id": "ID de l'application",
                "patent_title": "Titre",
                "Nb co-inventors": "Nb co-inventeurs locaux",
                "local inventors": "Inventeurs locaux",
                "inventors": "Inventeurs",
                "assignees": "Cessionnaires",
                "related_apps": "Applications liées",
            },
            inplace=True,
        )
        new_columns: list[str] = [
            "GUID",
            "Date de dépôt",
            "ID de l'application",
            "Titre",
            "Nb co-inventeurs locaux",
            "Inventeurs locaux",
            "Inventeurs",
            "Cessionnaires",
            "Applications liées",
        ]
        patents = patents.sort_values(by=["Date de dépôt"])
    else:
        patents.rename(
            columns={
                "publication_date": "Date de délivrance",
                "app_filing_date": "Date de dépôt",
                "guid": "GUID",
                "appl_id": "ID de l'application",
                "patent_title": "Titre",
                "Nb co-inventors": "Nb co-inventeurs locaux",
                "local inventors": "Inventeurs locaux",
                "inventors": "Inventeurs",
                "assignees": "Cessionnaires",
                "related_apps": "Applications liées",
            },
            inplace=True,
        )
        new_columns: list[str] = [
            "GUID",
            "Date de délivrance",
            "Date de dépôt",
            "ID de l'application",
            "Titre",
            "Nb co-inventeurs locaux",
            "Inventeurs locaux",
            "Inventeurs",
            "Cessionnaires",
            "Applications liées",
        ]
        patents = patents.sort_values(by=["Date de délivrance"])

    return patents[new_columns]


def _create_results_summary_df(
    reference_query: ReferenceQuery,
    publications_dfs_list_by_pub_type: list,
    patent_applications: pd.DataFrame,
    patents: pd.DataFrame,
) -> pd.DataFrame:
    """
    Create results summary dataframe

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        publications_dfs_list_by_pub_type (list): list of DataFrames with search results by type
        patent_applications (pd.DataFrame): patent application search results
        patents (pd.DataFrame): patent search results

    Returns: DataFrame with results summary

    """
    # Create results summary dataframe

    results: list = [
        None,
        "Nb d'auteur.e.s",
        "Année de début",
        "Année de fin",
    ]
    results += reference_query.publication_types
    results += ["Brevets US (en instance)", "Brevets US (délivrés)"]
    values: list = [
        None,
        len(reference_query.au_ids),
        reference_query.pub_year_first,
        reference_query.pub_year_last,
    ]
    if publications_dfs_list_by_pub_type:
        values += [
            0 if df.empty else len(df) for df in publications_dfs_list_by_pub_type
        ]
    else:
        values += [None] * len(reference_query.publication_types)
    values += [
        len(patent_applications),
        len(patents),
    ]
    co_authors: list = ["Conjointes", None, None, None]
    if publications_dfs_list_by_pub_type:
        co_authors += [
            None if df.empty else len(df[df["Nb co-auteurs locaux"] > 1])
            for df in publications_dfs_list_by_pub_type
        ]
    else:
        co_authors += [None] * len(reference_query.publication_types)

    joint_patent_applications_count: int = sum(
        row["Nb co-inventeurs locaux"] is not None
        and row["Nb co-inventeurs locaux"] > 1
        for _, row in patent_applications.iterrows()
    )
    joint_patents_count: int = sum(
        row["Nb co-inventeurs locaux"] is not None
        and row["Nb co-inventeurs locaux"] > 1
        for _, row in patents.iterrows()
    )
    co_authors += [joint_patent_applications_count, joint_patents_count]

    return pd.DataFrame([results, values, co_authors]).T


def write_reference_query_results_to_excel(
    reference_query: ReferenceQuery,
    publications_dfs_list_by_pub_type: list[pd.DataFrame],
    patents: pd.DataFrame,
    patent_applications: pd.DataFrame,
    author_profiles_by_ids: pd.DataFrame,
    author_profiles_by_names: pd.DataFrame,
) -> None:
    """
    Write publications search results to the output Excel file

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        publications_dfs_list_by_pub_type (list): list of DataFrames with search results by type
        patents (pd.DataFrame): patent application search results by filing date
        patent_applications (pd.DataFrame): patent search results by publication date
        author_profiles_by_ids (pd.DataFrame): author search results by ids
        author_profiles_by_names (pd.DataFrame): author search results by names

    Returns: None

    """

    # Create results summary dataframe
    results: pd.DataFrame = _create_results_summary_df(
        reference_query=reference_query,
        publications_dfs_list_by_pub_type=publications_dfs_list_by_pub_type,
        patent_applications=patent_applications,
        patents=patents,
    )

    # Write dataframes in separate sheets to the output Excel file
    with pd.ExcelWriter(reference_query.out_excel_file) as writer:
        # Results (first) sheet
        results.to_excel(writer, index=False, header=False, sheet_name="Résultats")

        # Write Scopus search results dataframes to separate sheets by publication type
        for df, pub_type in zip(
            publications_dfs_list_by_pub_type, reference_query.publication_types
        ):
            if not df.empty:
                _export_publications_df_to_excel_sheet(
                    writer=writer,
                    df=df,
                    sheet_name=pub_type,
                )

        # USPTO search result sheets
        if not patent_applications.empty:
            patent_applications.to_excel(
                writer,
                index=False,
                sheet_name="Brevets US (en instance)",
                freeze_panes=(1, 1),
            )
        if not patents.empty:
            patents.to_excel(
                writer,
                index=False,
                sheet_name="Brevets US (délivrés)",
                freeze_panes=(1, 1),
            )

        # Author profile sheets
        col: pd.Series = author_profiles_by_ids.pop("Période active")
        author_profiles_by_ids["Période active"] = col
        author_profiles_by_ids.to_excel(
            writer, index=False, sheet_name="Auteurs - Profils", freeze_panes=(1, 1)
        )
        author_profiles_by_names.to_excel(
            writer, index=False, sheet_name="Auteurs - Homonymes", freeze_panes=(1, 1)
        )
    print(
        "Résultats de la recherche sauvegardés "
        f"dans le fichier '{reference_query.out_excel_file}'"
    )

    # Attempt to adjust column widths in the output Excel file to reasonable values.
    # The solution is a hack because the auto_size/bestFit properties in
    # openpyxl.worksheet.dimensions.ColumnDimension() don't seem to work and the actual
    # column width sizing in Excel is system-dependant and a bit of a black box.
    workbook = load_workbook(reference_query.out_excel_file)
    col_width_max: int = 100
    for sheet_name in workbook.sheetnames:
        for i, col in enumerate(workbook[sheet_name].columns):
            # workbook[sheet_name].column_dimensions[col[0].column_letter].bestFit = True
            col_width: int = int(max(len(str(cell.value)) for cell in col) * 0.85)
            col_width_min: int = 18 if i == 0 else 10
            workbook[sheet_name].column_dimensions[col[0].column_letter].width = max(
                min(col_width_max, col_width), col_width_min
            )
    workbook.save(reference_query.out_excel_file)


def query_scopus_author_profiles_by_name(
    reference_query: ReferenceQuery,
    homonyms_only: bool = True,
) -> pd.DataFrame:
    """
    Fetch author profiles by name in the Scopus database. If homonyms_only is True,
    retain only author profiles with homonyms (multiple Scopus IDs for same name).

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        homonyms_only (bool): Include only author profiles with homonyms

    Returns: DataFrame with author search results

    """

    author_profiles_all = pd.DataFrame()
    for [lastname, firstname] in reference_query.au_names:
        query_string: str = f"AUTHLAST({lastname}) and AUTHFIRST({firstname})"
        author_profiles_from_name_search_results = AuthorSearch(
            query=query_string,
            refresh=reference_query.scopus_database_refresh,
            verbose=True,
        )
        if author_profiles_from_name_search_results.authors:
            author_profiles_from_name = pd.DataFrame(
                author_profiles_from_name_search_results.authors
            )
            author_profiles_from_name["eid"] = [
                au_id.split("-")[-1]
                for au_id in author_profiles_from_name.eid.to_list()
            ]
            (
                author_profiles_from_name["Start"],
                author_profiles_from_name["End"],
            ) = zip(
                *[
                    AuthorRetrieval(
                        author_id=au_id, refresh=reference_query.scopus_database_refresh
                    ).publication_range
                    for au_id in author_profiles_from_name.eid.to_list()
                ]
            )
            if not homonyms_only or author_profiles_from_name.shape[0] > 1:
                author_profiles_all = pd.concat(
                    [author_profiles_all, author_profiles_from_name],
                    ignore_index=True,
                )
                author_profiles_all.loc[len(author_profiles_all)] = [None] * len(
                    author_profiles_all.columns
                )
        elif not homonyms_only:
            print(
                f"[red]ERREUR: aucun résultat pour l'auteur.e '{lastname}, {firstname}' [/red]"
            )

    if not author_profiles_all.empty:
        author_profiles_all = _flag_matched_scopus_author_ids_and_affiliations(
            reference_query=reference_query, author_profiles=author_profiles_all
        )
        author_profiles_all = _reindex_author_profiles_df(df=author_profiles_all)

    return author_profiles_all


def query_scopus_author_profiles_by_id(reference_query: ReferenceQuery) -> pd.DataFrame:
    """

    Fetch author profiles from their IDs in the Scopus database

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object

    Returns: DataFrame with author profiles

    """

    author_profiles = []
    columns: list[str] = [
        "Nom de famille",
        "Prénom",
        "Affiliation",
        "Affiliation mère",
        "Période active",
    ]
    for i, [name, au_id] in enumerate(
        zip(reference_query.au_names, reference_query.au_ids)
    ):
        try:
            if au_id > 0:
                author = AuthorRetrieval(
                    author_id=au_id, refresh=reference_query.scopus_database_refresh
                )
                author_profiles.append(
                    [
                        author.surname,
                        author.given_name,
                        author.affiliation_current[0].__getattribute__(
                            "preferred_name"
                        ),
                        author.affiliation_current[0].__getattribute__(
                            "parent_preferred_name"
                        ),
                        author.publication_range,
                    ]
                )
            else:
                author_profiles.append([None] * len(columns))
        except ScopusException as e:
            vpn_required_str: str = (
                " ou tentative d'accès hors du réseau "
                "universitaire UdeS (VPN requis)"
                if i == 0
                else ""
            )
            print(
                f"[red]Erreur dans la recherche Scopus à la ligne {i + 2} "
                f"({name[0]}, {name[1]}) "
                f"du fichier {reference_query.in_excel_file}  - '{e}'[/red] - "
                f"Causes possibles: identifiant Scopus inconnu{vpn_required_str}!"
            )
            exit()

    # Create author profiles DataFrame, flag discrepancies between input and Scopus data
    author_profiles_by_ids: pd.DataFrame = pd.DataFrame()
    if author_profiles:
        author_profiles_by_ids = pd.DataFrame(author_profiles, columns=columns)
        author_profiles_by_ids.insert(
            loc=2,
            column="Erreurs",
            value=pd.Series(
                _check_author_name_correspondance(
                    reference_query=reference_query, authors=author_profiles_by_ids
                )
            ),
        )

    return author_profiles_by_ids


def query_scopus_publications(
    reference_query: ReferenceQuery,
) -> tuple[pd.DataFrame, list[list[int | None]]]:
    """
    Fetch publications for range of years in Scopus database for list of author IDs

    Scopus document type search terms:
      Article-ar / Abstract Report-ab / Book-bk / Book Chapter-ch / Conference Paper-cp /
      Conference Review-cr / Data Paper-dp / Editorial-ed / Erratum-er / Letter-le /
      Multimedia-mm / Note-no / Report-rp / Retracted-tb / Review-re / Short Survey-sh

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info

    Returns : DataFrame with publication search results (pd.DataFrame),
              list of publication type counts by author (list)
    """

    # Build Scopus document type search string
    pub_types_search_string: str = " OR ".join(
        [f"DOCTYPE ({s})" for s in reference_query.publication_type_codes]
    )

    # Loop through list of author IDs to fetch publications, count pub types by author
    publications = pd.DataFrame()
    pub_type_counts_by_author: list[list[int | None]] = []
    for au_id in reference_query.au_ids:
        if au_id > 0:
            query_str: str = (
                f"AU-ID ({au_id})"
                f" AND PUBYEAR > {reference_query.pub_year_first - 1}"
                f" AND PUBYEAR < {reference_query.pub_year_last + 1}"
                f" AND ({pub_types_search_string})"
            )
            try:
                query_results = ScopusSearch(
                    query=query_str,
                    refresh=reference_query.scopus_database_refresh,
                    verbose=True,
                )
            except ScopusException as e:
                print(
                    f"[red]Erreur dans la recherche Scopus pour l'identifiant {au_id}, "
                    f"causes possibles: identifiant inconnu ou tentative d'accès "
                    f"hors du réseau universitaire UdeS (VPN requis) - '{e}'[/red]"
                )
                exit()

            author_pubs = pd.DataFrame(query_results.results)
            pub_type_counts_by_author.append(
                _count_publications_by_type_in_df(
                    reference_query=reference_query, df=author_pubs
                )
            )
            publications = pd.concat([publications, author_pubs])
        else:
            pub_type_counts_by_author.append(
                [None] * len(reference_query.publication_type_codes)
            )
    pub_type_counts_by_author = np.transpose(pub_type_counts_by_author).tolist()

    if not publications.empty:
        publications = _add_coauthor_columns_and_clean_up_publications_df(
            publications, reference_query
        )

    return publications, pub_type_counts_by_author


def query_us_patents(
    reference_query: ReferenceQuery,
    applications: bool = True,
    application_ids_to_remove=None,
) -> tuple[pd.DataFrame, list, list]:
    """
    Query the USPTO database for patent applications or published patents
    for a list of authors over a range of years using the "patent_client" package

    See: https://patent-client.readthedocs.io/en/latest/user_guide/fulltext.html
         https://www.uspto.gov/patents/search/patent-public-search/quick-reference-guides

         USPTO database field codes for search over a range of years:
         - Applications: ((<first name>  NEAR2 <last name>).IN.) AND @AD>="<year0>0101"<="<year1>1231"
         - Patents: ((<first name> NEAR2 <last name>).IN.) AND @PD>="<year0>0101"<="<year1>1231"

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        applications (bool): Search filed applications if True, else search published patents if False
        application_ids_to_remove (list): List of application ids to remove from search results

    Returns : DataFrame with patent search results, list of patent application_ids

    """

    # Execute USPTO query (patent applications or delivered patents)
    if applications:
        print("En attente de la recherche de brevets en instance USPTO...", end="")
    else:
        print("En attente de la recherche de brevets USPTO...", end="")
    patents: pd.DataFrame = _query_uspto(
        reference_query=reference_query, applications=applications
    )
    print("terminé!")

    # Clean up USPTO search result dataframes
    application_ids: list[int] = []
    patent_counts_by_author: list[int | None] = [None] * len(reference_query.au_ids)
    if not patents.empty:
        # Simplify lists of inventors (names + country codes) and assignees (names)
        patents["inventors"] = patents["inventors"].apply(
            lambda inventors: [
                f"{inventor[0][1]} ({inventor[2][1]})" for inventor in inventors
            ]
        )
        patents["assignees"] = patents["assignees"].apply(
            lambda assignees: [assignee[2][1] for assignee in assignees]
        )

        # Remove dataframe rows with no Canadian inventors
        no_canadian_inventors: pd.Series = patents["inventors"].apply(
            lambda inventors: all("(CA)" not in inventor for inventor in inventors)
        )
        patents.drop(patents[no_canadian_inventors].index, inplace=True)

        # Add dataframe columns with lists and counts of local inventors
        patents["local inventors"] = patents["inventors"].apply(
            lambda inventors: [
                lastname
                for [lastname, firstname] in reference_query.au_names
                if any(
                    (
                        _to_lower_no_accents_no_hyphens(lastname)
                        in _to_lower_no_accents_no_hyphens(inventor)
                    )
                    and (
                        _to_lower_no_accents_no_hyphens(firstname)
                        in _to_lower_no_accents_no_hyphens(inventor)
                    )
                    for inventor in inventors
                )
            ]
        )
        patents["Nb co-inventors"] = patents["local inventors"].apply(
            lambda inventors: len(inventors) if len(inventors) > 1 else None
        )

        # Remove dataframe rows with no local inventors
        no_local_inventors: pd.Series = patents["local inventors"].apply(
            lambda inventors: not inventors
        )
        patents.drop(patents[no_local_inventors].index, inplace=True)

        # Compile list of patent/application ids. Remove applications for which patents
        # have been delivered(patent applications having same IDs as delivered patents)
        application_ids: list = patents["appl_id"].to_list()
        if applications and application_ids_to_remove:
            patents.drop(
                patents[
                    [
                        application_id in application_ids_to_remove
                        for application_id in application_ids
                    ]
                ].index,
                inplace=True,
            )

        # Reorder columns, change names to French, sort by date
        patents = _reformat_uspto_search_results(
            patents=patents, applications=applications
        )

        # Tabulate number of patents or patent applications per author
        patent_counts_by_author = _tabulate_patents_per_author(
            reference_query=reference_query, patents=patents
        )

    return patents, application_ids, patent_counts_by_author


def query_epo_patents(reference_query: ReferenceQuery) -> None:
    """

    To connect to the European Patent Office’s Open Patent Services, an API key is
    required, see:
    - https://patent-client.readthedocs.io/en/stable/getting_started.html
    - https://www.epo.org/en/searching-for-patents/data/web-services/ops

    Espacenet query string example (https://worldwide.espacenet.com/patent/search):
    (in=("charette" prox/distance<1 "paul") OR in=("hunter" prox/distance<1 "ian")) AND pd within "1990,2020"

    """

    def inventor_query_str(inventor: list[str]) -> str:
        return f'in=("{inventor[1]}" prox/distance<1 "{inventor[0]}")'

    def build_epo_patent_query_string() -> str:
        s: str = "("
        s += inventor_query_str(reference_query.au_names[0])
        for name in reference_query.au_names[1:]:
            s += " OR "
            s += inventor_query_str(name)
        s += (
            f") AND pd within "
            f'"{reference_query.pub_year_first},{reference_query.pub_year_last}"'
        )
        return s

    max_results: int = 500
    query_str: str = build_epo_patent_query_string()

    results = Inpadoc.objects.filter(cql_query='inventor="Charette Paul"')
    n = len(results)
    print("EPO search done!")


def query_publications_and_patents(reference_query: ReferenceQuery) -> None:
    """
    Search for publications in Scopus and patents in the USPTO database
    for a list of authors over a range of years

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info

    Returns: None

    """

    # Console banner
    print(
        "Recherche de publications et brevets US pour la période "
        f"{reference_query.pub_year_first}-{reference_query.pub_year_last}"
    )

    # Fetch author profiles corresponding to user-supplied Scopus IDs, check they match
    # the user-supplied names, flag any inconsistencies in the "Erreurs" column
    author_profiles_by_ids: pd.DataFrame = query_scopus_author_profiles_by_id(
        reference_query=reference_query
    )

    # Fetch publications by type in Scopus database, count publication types by author
    publications_all: pd.DataFrame
    pub_type_counts_by_author: list[list[int | None]]
    publications_all, pub_type_counts_by_author = query_scopus_publications(
        reference_query=reference_query
    )

    # Loop to parse publications by type into separate dataframes, store dfs in a list
    publications_dfs_list_by_pub_type: list[pd.DataFrame] = []
    if not publications_all.empty:
        for [pub_type, pub_code, pub_counts] in zip(
            reference_query.publication_types,
            reference_query.publication_type_codes,
            pub_type_counts_by_author,
        ):
            # Extract "pub_type" publications into a dataframe, add dataframe to list
            df: pd.DataFrame = publications_all[publications_all["subtype"] == pub_code]
            publications_dfs_list_by_pub_type.append(df)
            print(f"{pub_type}: {len(df)}")

            # Add "pub_type" publication counts to the author profiles
            if len(df) > 0:
                author_profiles_by_ids[pub_type] = pub_counts

    # Fetch US applications and published patents into separate dataframes, if required
    patents: pd.DataFrame = pd.DataFrame()
    patent_applications: pd.DataFrame = pd.DataFrame()
    if reference_query.include_patents:
        patent_application_ids: list
        patent_counts_by_author: list
        patents, patent_application_ids, patent_counts_by_author = query_us_patents(
            reference_query=reference_query, applications=False
        )
        print("Brevets US (délivrés): ", len(patents))
        patent_application_counts_by_author: list
        patent_applications, _, patent_application_counts_by_author = query_us_patents(
            reference_query=reference_query,
            applications=True,
            application_ids_to_remove=patent_application_ids,
        )
        print("Brevets US (en instance): ", len(patent_applications))

        # Add patent application and published patent counts to the author profiles
        author_profiles_by_ids["Brevets US (en instance)"] = (
            patent_application_counts_by_author
        )
        author_profiles_by_ids["Brevets US (délivrés)"] = patent_counts_by_author

    # Fetch Scopus author profiles corresponding to user-supplied names, check for
    # author names with multiple Scopus IDs ("homonyms"), load into dataframe
    author_profiles_by_name: pd.DataFrame = query_scopus_author_profiles_by_name(
        reference_query=reference_query,
        homonyms_only=True,
    )

    # Write results to output Excel file
    write_reference_query_results_to_excel(
        reference_query=reference_query,
        publications_dfs_list_by_pub_type=publications_dfs_list_by_pub_type,
        patents=patents,
        patent_applications=patent_applications,
        author_profiles_by_ids=author_profiles_by_ids,
        author_profiles_by_names=author_profiles_by_name,
    )


def query_author_profiles(reference_query: ReferenceQuery) -> None:
    """
    Query Scopus for a list of author profiles by name

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info

    Returns: None

    """

    print("Recherche de profils d'auteur.e.s par nom dans la base de données Scopus")
    author_profiles_by_name: pd.DataFrame = query_scopus_author_profiles_by_name(
        reference_query=reference_query,
        homonyms_only=False,
    )
    author_profiles_by_name.rename(
        columns={
            "eid": "ID Scopus",
        },
        inplace=True,
    )
    with pd.ExcelWriter(reference_query.out_excel_file) as writer:
        author_profiles_by_name.to_excel(writer, index=False, sheet_name="Profils")
    print(
        "Résultats de la recherche sauvegardés "
        f"dans le fichier '{reference_query.out_excel_file}'"
    )


def run_reference_search(reference_query: ReferenceQuery, search_type: str) -> None:
    """
     For a list of author names and range of years, search either for:
        - references (publications in Scopus, patents in the USPTO database)
          OR
        - author profiles in the Scopus database

     Args:
         reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
         search_type (str): Type of search ("Publications" or "Profils")

    Returns: None

    """

    # Console info starting messages
    python_version: str = (
        f"{str(sys.version_info.major)}"
        f".{str(sys.version_info.minor)}"
        f".{str(sys.version_info.micro)}"
    )
    print(f"{Path(__file__).stem} {__version__} " f"(running python {python_version})")
    print(
        f"Nombre d'auteur.e.s dans le fichier '{reference_query.in_excel_file}': "
        f"{len(reference_query.au_ids)}"
    )

    # Init Scopus API
    pybliometrics.scopus.init()

    # Run the query
    if search_type == "Publications":
        query_publications_and_patents(reference_query=reference_query)
    elif search_type == "Profils":
        query_author_profiles(reference_query=reference_query)
    else:
        print(
            f"[red]ERREUR: '{search_type}' est un type de recherche invalide, "
            "doit être 'Author name' ou 'Scopus ID'[/red]"
        )


def main():
    # Load search parameters from toml file
    toml_dict: dict = toml.load("pyRefSearchUdeS.toml")

    # Define input/output Excel file names
    in_excel_file: Path = Path(toml_dict["in_excel_file"])
    out_excel_file: Path = (
        Path(
            f"{in_excel_file.stem}_publications_"
            f"{toml_dict['pub_year_first']}-{toml_dict['pub_year_last']}"
            f"{in_excel_file.suffix}"
        )
        if toml_dict["search_type"] == "Publications"
        else Path(f"{in_excel_file.stem}_profils" f"{in_excel_file.suffix}")
    )

    # Define ReferenceQuery Class object containing the query parameters
    reference_query: ReferenceQuery = ReferenceQuery(
        in_excel_file=in_excel_file,
        in_excel_file_author_sheet=toml_dict["in_excel_file_author_sheet"],
        out_excel_file=out_excel_file,
        pub_year_first=toml_dict["pub_year_first"],
        pub_year_last=toml_dict["pub_year_last"],
        publication_types=toml_dict["publication_types"],
        local_affiliations=toml_dict["local_affiliations"],
        scopus_database_refresh=toml_dict["scopus_database_refresh"],
        include_patents=toml_dict.get("include_patents", True),
    )

    # query_epo_patents(reference_query)

    # Run the bibliographic search!
    run_reference_search(
        reference_query=reference_query,
        search_type=toml_dict["search_type"],
    )


if __name__ == "__main__":
    main()
    print("pyScopus terminé!")
