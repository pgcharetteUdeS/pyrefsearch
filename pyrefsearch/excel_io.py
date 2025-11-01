"""excel_io.py

Excel file I/O utilities

"""

__all__ = [
    "load_espacenet_search_results_from_excel_file",
    "write_espacenet_search_results_to_excel_file",
    "write_reference_query_results_to_excel_file",
]

import ast
import datetime
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from pathlib import Path
import re
import sys
import time

from referencequery import ReferenceQuery
from utils import console


def _export_publications_df_to_excel_sheet(
    writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str
) -> list:
    """
    Write selected set of publication dataframe columns to Excel sheet

    Args:
        writer (pd.ExcelWriter): openpyxl writer object
        df (pd.DataFrame): articles dataframe
        sheet_name (str): Excel file sheet name

    Returns: list of column names to write to the Excel file

    """

    column_names: list[str] = [
        "Titre",
        "Date",
        "Auteurs locaux",
        "Collab interne",
        "Auteurs",
        "Publication",
        "Volume",
        "DOI",
    ]
    if not df.empty:
        df_copy: pd.DataFrame = df.rename(
            columns={
                "coverDate": "Date",
                "title": "Titre",
                "Auteurs locaux": "Auteurs locaux",
                "Collab interne": "Collab interne",
                "author_names": "Auteurs",
                "publicationName": "Publication",
                "volume": "Volume",
                "doi": "DOI",
            },
        ).copy()
        df_copy[column_names].to_excel(
            writer, index=False, sheet_name=sheet_name, freeze_panes=(1, 1)
        )
    return column_names


def _create_results_summary_df(
    reference_query: ReferenceQuery,
    publications_dfs_list_by_pub_type: list,
    uspto_patent_applications: pd.DataFrame = pd.DataFrame([]),
    uspto_patents: pd.DataFrame = pd.DataFrame([]),
    inpadoc_patent_applications=pd.DataFrame([]),
    inpadoc_patents: pd.DataFrame = pd.DataFrame([]),
) -> pd.DataFrame:
    """
    Create results summary dataframe
    - "results" (first column): search information labels and bibliographic item names
    - "values" (second column): search information and result counts
    - "co_authors" (third column): number of local co-authors for each publication type

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        publications_dfs_list_by_pub_type (list): list of DataFrames with search results by type
        uspto_patent_applications (pd.DataFrame): uspto patent application search results
        uspto_patents (pd.DataFrame): uspto patent search results
        inpadoc_patent_applications (pd.DataFrame): inpadoc patent search results
        inpadoc_patents (pd.DataFrame): inpadoc patent search results

    Returns: DataFrame with results summary

    """

    # Initialize 3 columns contents
    results: list = [
        f"Recherche dans {reference_query.publications_search_database}",
        "Nb d'auteur.e.s",
        "Année de début",
        "Année de fin",
    ]
    values: list = [
        None,
        len(reference_query.scopus_ids),
        reference_query.pub_year_first,
        reference_query.pub_year_last,
    ]
    co_authors: list = ["Conjointes", None, None, None]

    # Publications search results
    results += reference_query.publication_types
    if publications_dfs_list_by_pub_type:
        values += [
            0 if df.empty else len(df) for df in publications_dfs_list_by_pub_type
        ]
        co_authors += [
            None if df.empty else len(df[df["Collab interne"] > 1])
            for df in publications_dfs_list_by_pub_type
        ]
    else:
        values += [None] * len(reference_query.publication_types)
        co_authors += [None] * len(reference_query.publication_types)

    # USPTO search results
    if not uspto_patents.empty or not uspto_patent_applications.empty:
        results += ["Brevets USPTO (en instance)", "Brevets USPTO (délivrés)"]
        values += [
            len(uspto_patent_applications),
            len(uspto_patents),
        ]
        uspto_joint_patent_applications_count: int = sum(
            row["Nb co-inventeurs locaux"] is not None
            and row["Nb co-inventeurs locaux"] > 1
            for _, row in uspto_patent_applications.iterrows()
        )
        uspto_joint_patents_count: int = sum(
            row["Nb co-inventeurs locaux"] is not None
            and row["Nb co-inventeurs locaux"] > 1
            for _, row in uspto_patents.iterrows()
        )
        co_authors += [uspto_joint_patent_applications_count, uspto_joint_patents_count]

    # INPADOC search results
    if not inpadoc_patents.empty or not inpadoc_patent_applications.empty:
        results += ["Brevets INPADOC (en instance)", "Brevets INPADOC (délivrés)"]
        values += [
            len(inpadoc_patent_applications),
            len(inpadoc_patents),
        ]
        inpadoc_patent_applications_count: int = sum(
            row["Nb co-inventeurs locaux"] is not None
            and row["Nb co-inventeurs locaux"] > 1
            for _, row in inpadoc_patent_applications.iterrows()
        )
        inpadoc_patents_count: int = sum(
            row["Nb co-inventeurs locaux"] is not None
            and row["Nb co-inventeurs locaux"] > 1
            for _, row in inpadoc_patents.iterrows()
        )
        co_authors += [inpadoc_patent_applications_count, inpadoc_patents_count]

    return pd.DataFrame([results, values, co_authors]).T


def _center_column_by_index(worksheet: Worksheet, i: int):
    for row in worksheet:
        cell = row[i]
        cell.alignment = Alignment(horizontal="center")


def _add_totals_formulae_to_sheet(
    worksheet: Worksheet, n: int, column_names: list
) -> None:
    """
    Add total and % totals at the end of an Excel sheet in columns A, "Collab interne"
    and "Collab externe". Format the data in the columns.

    Args
        worksheet (Worksheet): worksheet to which to add the totals
        n (int): number of data rows in the worksheet
        column_names (list): list of column names in the worksheet

    Returns: None

    """

    # Add summing formula to first column
    worksheet[f"A{n + 2}"] = "NOMBRE TOTAL"
    worksheet[f"A{n + 2}"].border = Border(top=Side(style="thin"))
    worksheet[f"A{n + 2}"].alignment = Alignment(horizontal="right")
    worksheet[f"A{n + 3}"] = f"=COUNTA(A2:A{n + 1})"

    # Add % sum formula to column "Collab interne"
    col_name = "Collab interne"
    col = get_column_letter(column_names.index(col_name) + 1)
    worksheet[f"{col}1"].alignment = Alignment(wrapText=True)
    worksheet[f"{col}{n + 2}"] = "% DU TOTAL"
    worksheet[f"{col}{n + 2}"].border = Border(top=Side(style="thin"))
    worksheet[f"{col}{n + 2}"].alignment = Alignment(horizontal="right")
    worksheet[f"{col}{n + 3}"] = f"=ROUND(COUNTA({col}2:{col}{n + 1})/A{n + 3}*100, 1)"
    _center_column_by_index(worksheet=worksheet, i=column_names.index(col_name))


def write_reference_query_results_to_excel_file(
    reference_query: ReferenceQuery,
    publications: pd.DataFrame,
    pub_type_counts_by_author: list,
    author_profiles_by_name: pd.DataFrame,
    uspto_patents: pd.DataFrame,
    uspto_patent_applications: pd.DataFrame,
    inpadoc_patents: pd.DataFrame,
    inpadoc_patent_applications: pd.DataFrame,
    # scopus_author_profiles_by_ids: pd.DataFrame,
    publications_diff: bool = False,
    publications_previous_filename: Path = Path(""),
) -> Path:
    """
    Write publications search results to the output Excel file

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info
        publications (pd.DataFrame): dataFrames publications found in Scopus
        pub_type_counts_by_author (list): lists of publication type by author from Scopus
        author_profiles_by_name (pd.DataFrame): Scopus author search results by names
        uspto_patents (pd.DataFrame): USPTO patent application search results
        uspto_patent_applications (pd.DataFrame): USPTO patent search results
        inpadoc_patents (pd.DataFrame): INPADOC patent search result
        inpadoc_patent_applications (pd.DataFrame): INPADOC patent search results
        scopus_author_profiles_by_ids (pd.DataFrame): Scopus author search results by ids
        publications_diff (bool): True of this a Scopus differential request
        publications_previous_filename (Path): Path to the Excel file with results from previous month

    Returns (Path): Excel output filename

    """

    # Loop to parse publications by type into separate dataframes, store dfs in a list
    publications_dfs_list_by_pub_type: list[pd.DataFrame] = []
    if not publications.empty:
        for [pub_type, pub_code, pub_counts] in zip(
            reference_query.publication_types,
            reference_query.publication_type_codes,
            pub_type_counts_by_author,
        ):
            # Extract "pub_type" publications into a dataframe, add dataframe to list
            df: pd.DataFrame = pd.DataFrame(
                publications[publications["subtype"] == pub_code]
            )
            publications_dfs_list_by_pub_type.append(df)
            console.print(f"{pub_type}: {len(df)}")

            # Add "pub_type" publication counts to the author profiles
            """
            if len(df) > 0:
                scopus_author_profiles_by_ids[pub_type] = pub_counts
            """

    # Create results summary dataframe
    results: pd.DataFrame = _create_results_summary_df(
        reference_query=reference_query,
        publications_dfs_list_by_pub_type=publications_dfs_list_by_pub_type,
        uspto_patent_applications=uspto_patent_applications,
        uspto_patents=uspto_patents,
        inpadoc_patent_applications=inpadoc_patent_applications,
        inpadoc_patents=inpadoc_patents,
    )

    # Write dataframes in separate sheets to the output Excel file
    out_excel_filename: Path = (
        reference_query.out_excel_file.with_stem(
            f"{reference_query.out_excel_file.stem}_SCOPUS_DIFF_"
            f"{publications_previous_filename.stem[-len('YYYY-MM-DD'):]}"
        )
        if publications_diff
        else reference_query.out_excel_file
    )
    with pd.ExcelWriter(out_excel_filename) as writer:
        # Results (first) sheet
        results.to_excel(writer, index=False, header=False, sheet_name="Résultats")

        # Write Scopus search results dataframes to separate sheets by publication type
        for df, pub_type in zip(
            publications_dfs_list_by_pub_type, reference_query.publication_types
        ):
            if not df.empty:
                # Write columns
                column_names: list = _export_publications_df_to_excel_sheet(
                    writer=writer,
                    df=df,
                    sheet_name=pub_type,
                )
                _add_totals_formulae_to_sheet(
                    worksheet=writer.sheets[pub_type],
                    n=len(df),
                    column_names=column_names,
                )
                _center_column_by_index(
                    worksheet=writer.sheets[pub_type],
                    i=column_names.index("Auteurs locaux"),
                )

        if not publications_diff:
            # Write all publication search results to a single sheet
            publications.to_excel(
                writer,
                index=False,
                sheet_name=f"Résultats complets {reference_query.publications_search_database}",
                freeze_panes=(1, 1),
            )

            # USPTO search result sheets
            if not uspto_patent_applications.empty:
                uspto_patent_applications.to_excel(
                    writer,
                    index=False,
                    sheet_name="Brevets US (en instance)",
                    freeze_panes=(1, 1),
                )
            if not uspto_patents.empty:
                uspto_patents.to_excel(
                    writer,
                    index=False,
                    sheet_name="Brevets US (délivrés)",
                    freeze_panes=(1, 1),
                )

            # INPADOC search result sheets
            if not inpadoc_patent_applications.empty:
                inpadoc_patent_applications.to_excel(
                    writer,
                    index=False,
                    sheet_name="Brevets INPADOC (en instance)",
                    freeze_panes=(1, 1),
                )
            if not inpadoc_patents.empty:
                inpadoc_patents.to_excel(
                    writer,
                    index=False,
                    sheet_name="Brevets INPADOC (délivrés)",
                    freeze_panes=(1, 1),
                )

            # Author profile sheets
            """
            col: pd.Series = scopus_author_profiles_by_ids.pop("Période active")
            scopus_author_profiles_by_ids["Période active"] = col
            scopus_author_profiles_by_ids.to_excel(
                writer, index=False, sheet_name="Auteurs - Profils", freeze_panes=(1, 1)
            )
            """
            author_profiles_by_name.to_excel(
                writer,
                index=False,
                sheet_name="Auteurs - Homonymes",
                freeze_panes=(1, 1),
            )

        """
        else:
            # Author profile sheets
            author_profiles_by_ids_minimal: pd.DataFrame = (
                scopus_author_profiles_by_ids[["Nom de famille", "Prénom"]].copy()
            )
            author_profiles_by_ids_minimal.to_excel(
                writer, index=False, sheet_name="Auteurs", freeze_panes=(1, 1)
            )
        """

    console.print(
        "Résultats de la recherche sauvegardés "
        f"dans le fichier '{out_excel_filename}'",
        soft_wrap=True,
    )

    # Attempt to adjust column widths in the output Excel file to reasonable values.
    # The solution is a hack because the auto_size/bestFit properties in
    # openpyxl.worksheet.dimensions.ColumnDimension() don't seem to work and the actual
    # column width sizing in Excel is system-dependant and a bit of a black box.
    workbook = load_workbook(out_excel_filename)
    col_width_max: int = 100
    for sheet_name in workbook.sheetnames:
        for i, col in enumerate(workbook[sheet_name].columns):
            # workbook[sheet_name].column_dimensions[col[0].column_letter].bestFit = True
            col_width: int = int(max(len(str(cell.value)) for cell in col) * 0.85)
            col_width_min: int = 20 if i == 0 else 10
            workbook[sheet_name].column_dimensions[col[0].column_letter].width = max(
                min(col_width_max, col_width), col_width_min
            )
    workbook.save(out_excel_filename)

    # Return Excel output filename
    return out_excel_filename


def load_espacenet_search_results_from_excel_file(
    reference_query: ReferenceQuery,
) -> pd.DataFrame:
    """
    Load previous espacenet search results from Excel file, where search date is in the
    file name <filename>YYYYMMDD.xlsx

    Args:
        reference_query (ReferenceQuery): ReferenceQuery Class object containing query info

    Returns: DataFrame with espacenet search results

    """

    # Extract date from file name, show warning on console if file is older than 30 days
    if not (
        match := re.search(
            r"(\d{8}).xlsx",
            reference_query.espacenet_patent_search_results_file,
        )
    ):
        console.print(
            "[red]Impossible d'extraire la date du fichier de résultats de recherche"
            f" '{reference_query.espacenet_patent_search_results_file}' "
            "qui doit être en format '<filename>YYYYMMDD.xlsx'![/red]",
            soft_wrap=True,
        )
        sys.exit()

    file_date = datetime.datetime.strptime(match[1], "%Y%m%d").date()
    if datetime.date.today() - file_date >= timedelta(days=30):
        console.print(
            "[yellow]WARNING: Les données dans le fichier "
            f"'{reference_query.espacenet_patent_search_results_file}' "
            "ont plus de 30 jours![/yellow]",
            soft_wrap=True,
        )

    # Load data from Excel file
    patent_families: pd.DataFrame = pd.read_excel(
        reference_query.data_dir
        / Path(reference_query.espacenet_patent_search_results_file)
    )

    def parse_list_field(value):
        """
        Attempts to parse a string representation of a list.
        First, it tries using ast.literal_eval. If that fails,
        it falls back to regex-based extraction.
        """
        try:
            parsed_value = ast.literal_eval(value)
            if isinstance(parsed_value, list):
                return parsed_value
        except (ValueError, SyntaxError):
            # Fall back to extracting items between apostrophes.
            return re.findall(r"'([^']+)'", value)
        return []

    # Reformat inventors and applicants columns into proper lists using the robust parser
    patent_families["Inventeurs"] = patent_families["Inventeurs"].apply(
        parse_list_field
    )
    patent_families["Cessionnaires"] = patent_families["Cessionnaires"].apply(
        parse_list_field
    )

    return patent_families


def write_espacenet_search_results_to_excel_file(
    reference_query: ReferenceQuery, patent_families: pd.DataFrame
) -> None:
    # Write dataframe of all patent results to an Excel file
    with pd.ExcelWriter(
        reference_query.data_dir
        / Path(f"espacenet_search_results_{time.strftime('%Y%m%d')}.xlsx")
    ) as writer:
        patent_families.to_excel(
            writer,
            index=False,
            header=True,
            sheet_name="Recherche par inventeurs",
            freeze_panes=(1, 1),
        )
    fname: Path = reference_query.data_dir / Path(
        f"espacenet_search_results_{time.strftime('%Y%m%d')}.xlsx"
    )
    console.print(
        f"Résultats de la recherche dans espacenet sauvegardés dans le fichier '{fname}'",
        soft_wrap=True,
    )
